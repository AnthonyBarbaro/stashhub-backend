#!/usr/bin/env python3

"""
BrandInventoryGUIAdvanced.py

GUI that:
1) Lets you pick a folder of CSVs and an output folder (loaded from/saved to config.txt).
2) Loads brand names found in the CSV 'Brand' column (lowercased and trimmed).
3) Filters data to the selected brand(s) & splits them into "Available" (>2) and "Unavailable" (<=2),
   generating one XLSX per brand with advanced Excel formatting.
4) Uploads each brand’s XLSX to a date-based folder in Google Drive: 
     INVENTORY -> <YYYY-MM-DD> -> <brandName>  (folder is made public).
5) Sends an HTML email with each brand's public Drive folder link to the specified recipients.

Packages needed:
 - pandas, openpyxl
 - google-auth, google-auth-oauthlib, google-api-python-client
 - credentials.json for Google OAuth (Drive + Gmail)
 - token_drive.json, token_gmail.json are created automatically after first login.
 - config.txt (optional; stores your input/output folder paths).
"""

import os
import re
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import traceback
from datetime import datetime
import subprocess

# For Excel formatting
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Google API imports
import google.auth.transport.requests
from google_auth_oauthlib.flow import InstalledAppFlow
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

# ----------------- CONFIG -----------------

# File where we store the chosen input & output dirs.
CONFIG_FILE = "config.txt"

# Google Drive parent folder name
DRIVE_PARENT_FOLDER_NAME = "INVENTORY"

# OAuth credential files
CREDENTIALS_FILE = "credentials.json"
TOKEN_DRIVE_FILE = "token_drive.json"
TOKEN_GMAIL_FILE = "token_gmail.json"

# Google Drive API Scopes
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.file"]

# Gmail API Scopes
GMAIL_SCOPES = ["https://www.googleapis.com/auth/gmail.send"]

# Required CSV columns + optional
REQUIRED_COLUMNS = ["Available", "Product", "Brand"]
OPTIONAL_COLUMNS = ["Category", "Cost"]

# We'll consider Available <= 2 => "Unavailable"
MAX_AVAIL_FOR_UNAVAILABLE = 2

# ----------------------------------------------------------------------
#                  CONFIG.TXT load/save
# ----------------------------------------------------------------------
def load_config():
    """
    Reads the first two lines of config.txt:
        1) input_dir
        2) output_dir
    If missing or invalid, returns (None, None).
    """
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                lines = f.read().strip().split("\n")
                if len(lines) >= 2:
                    input_dir = lines[0].strip()
                    output_dir = lines[1].strip()
                    return input_dir, output_dir
        except:
            pass
    return None, None

def save_config(input_dir, output_dir):
    """
    Writes input_dir and output_dir to config.txt so next run loads them automatically.
    """
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            f.write(input_dir + "\n")
            f.write(output_dir + "\n")
    except Exception as e:
        print(f"[ERROR] Could not write config.txt: {e}")

# ----------------------------------------------------------------------
#                  GOOGLE DRIVE / GMAIL AUTH
# ----------------------------------------------------------------------
def drive_authenticate():
    """Authenticate & build the Google Drive service using OAuth."""
    creds = None
    if os.path.exists(TOKEN_DRIVE_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_DRIVE_FILE, DRIVE_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(google.auth.transport.requests.Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, DRIVE_SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_DRIVE_FILE, "w") as token:
            token.write(creds.to_json())
    return build("drive", "v3", credentials=creds)

def gmail_authenticate():
    """Authenticate with Gmail API (OAuth) and return a service object."""
    creds = None
    if os.path.exists(TOKEN_GMAIL_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_GMAIL_FILE, GMAIL_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(google.auth.transport.requests.Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, GMAIL_SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_GMAIL_FILE, "w") as f:
            f.write(creds.to_json())
    return build("gmail", "v1", credentials=creds)

def make_folder_public(drive_service, folder_id):
    """Make the given folder ID publicly viewable."""
    permission = {"type": "anyone", "role": "reader"}
    drive_service.permissions().create(fileId=folder_id, body=permission).execute()

def find_or_create_folder(drive_service, folder_name, parent_id=None, make_public=False):
    """
    Find or create a folder named folder_name under parent_id.
    If newly created and make_public=True, sets public read permission.
    Returns folder_id or None on error.
    """
    from googleapiclient.errors import HttpError
    folder_name_escaped = folder_name.replace("'", "\\'")
    q = f"mimeType='application/vnd.google-apps.folder' and name='{folder_name_escaped}'"
    if parent_id:
        q += f" and '{parent_id}' in parents"

    try:
        res = drive_service.files().list(q=q, spaces="drive", fields="files(id, name)").execute()
        folders = res.get("files", [])
    except HttpError as e:
        print(f"[ERROR] find_or_create_folder: {e}")
        return None

    if folders:
        return folders[0]["id"]

    meta = {
        "name": folder_name,
        "mimeType": "application/vnd.google-apps.folder",
    }
    if parent_id:
        meta["parents"] = [parent_id]

    new_folder = drive_service.files().create(body=meta, fields="id").execute()
    fid = new_folder.get("id")
    print(f"[INFO] Created new folder '{folder_name}' (ID: {fid})")

    if make_public:
        try:
            make_folder_public(drive_service, fid)
        except Exception as e:
            print(f"[ERROR] Could not make folder public: {e}")

    return fid

def upload_file_to_drive(drive_service, file_path, parent_id):
    """Upload a local file to the given parent folder ID. Return the uploaded file ID."""
    file_name = os.path.basename(file_path)
    meta = {"name": file_name, "parents": [parent_id]}
    media = MediaFileUpload(file_path, resumable=True)
    uploaded = drive_service.files().create(body=meta, media_body=media, fields="id").execute()
    return uploaded.get("id")

def send_email_with_gmail_html(subject, html_body, recipients):
    """
    Sends an HTML email via the Gmail API. 
    recipients can be a list or a single comma-separated string.
    """
    import base64
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    if isinstance(recipients, str):
        recipients = [r.strip() for r in recipients.split(",") if r.strip()]

    service = gmail_authenticate()

    msg = MIMEMultipart("alternative")
    msg["From"] = "me"
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = subject

    part_html = MIMEText(html_body, "html")
    msg.attach(part_html)

    raw_message = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    body = {"raw": raw_message}

    sent = service.users().messages().send(userId="me", body=body).execute()
    print(f"[GMAIL] Email sent! ID: {sent['id']} | Subject: {subject}")

# ----------------- EXCEL FORMATTING -----------------
def advanced_format_excel(xlsx_path):
    """Freeze top row, bold grey headers, auto-fit columns, group by 'Category'."""
    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        # Freeze row 1
        ws.freeze_panes = "A2"

        # Header style
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill

        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = cell.value
                if val is not None:
                    length = len(str(val))
                    if length > max_len:
                        max_len = length
            ws.column_dimensions[col_letter].width = max_len + 3

        # Insert grouping rows for 'Category'
        category_index = None
        for i, cell in enumerate(ws[1], start=1):
            if (cell.value or "").lower() == "category":
                category_index = i
                break
        if category_index:
            rows_data = list(ws.iter_rows(min_row=2, values_only=True))
            if rows_data:
                current_cat = None
                insert_positions = []
                row_num = 2
                for row_vals in rows_data:
                    cat_val = row_vals[category_index - 1]
                    if cat_val != current_cat:
                        if current_cat is not None:
                            insert_positions.append(row_num)
                        current_cat = cat_val
                    row_num += 1
                # Insert at the very top
                insert_positions.insert(0, 2)

                cat_font = Font(bold=True, size=14)
                cat_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

                # We'll also need the cat value
                row_num = 2
                cat_list = []
                cur_cat = None
                for row_vals in rows_data:
                    cat_val = row_vals[category_index - 1]
                    if cat_val != cur_cat:
                        cat_list.append((row_num, cat_val))
                        cur_cat = cat_val
                    row_num += 1

                # Insert from bottom to top
                for (pos, cat_val) in reversed(cat_list):
                    ws.insert_rows(pos, 1)
                    c = ws.cell(row=pos, column=1)
                    c.value = str(cat_val)
                    c.font = cat_font
                    c.fill = cat_fill
                    c.alignment = Alignment(horizontal='center', vertical='center')
    wb.save(xlsx_path)

def extract_strain_type(product_name):
    """Optional: parse 'S', 'H', 'I' from product name, if you want to track strain."""
    if not isinstance(product_name, str):
        return ""
    text = " " + product_name.upper() + " "
    if re.search(r"\bS\b", text):
        return "S"
    if re.search(r"\bH\b", text):
        return "H"
    if re.search(r"\bI\b", text):
        return "I"
    return ""

# ----------------- CSV -> XLSX: Avail + Unavail -----------------
def generate_brand_reports(csv_path, out_dir, selected_brands):
    """
    Splits CSV rows into:
      - Available: Available>2
      - Unavailable: Available<=2
    Then for each brand in "available", produce one XLSX with 2 sheets:
      - "Available" (the brand’s rows)
      - "Unavailable" (the brand’s rows from the unavailable set, if any)
    Returns { brand_lower: [list_of_xlsx_paths] } for each brand found.
    """
    try:
        df = pd.read_csv(csv_path)
    except Exception as e:
        print(f"[ERROR] reading {csv_path}: {e}")
        return {}

    # Keep relevant columns only
    keep_cols = [c for c in REQUIRED_COLUMNS + OPTIONAL_COLUMNS if c in df.columns]
    if not any(c in df.columns for c in REQUIRED_COLUMNS):
        print(f"[WARN] '{csv_path}' is missing required columns {REQUIRED_COLUMNS}. Skipping.")
        return {}

    df = df[keep_cols]

    # Remove "sample"/"promo" lines
    if "Product" in df.columns:
        df = df[~df["Product"].str.contains(r"(?i)\bsample\b|\bpromo\b", na=False)]

    # Split into available/unavailable
    unavailable_df = df[df["Available"] <= MAX_AVAIL_FOR_UNAVAILABLE]
    available_df = df[df["Available"] > MAX_AVAIL_FOR_UNAVAILABLE]

    if "Brand" not in available_df.columns or available_df.empty:
        print(f"[INFO] No brand data or empty after filtering in '{csv_path}'")
        return {}

    # Lowercase + strip brand for consistent matching
    available_df["Brand"] = available_df["Brand"].astype(str).str.strip().str.lower()

    # If user selected brand(s), also convert them to lowercase
    if selected_brands:
        # Turn each user brand into a lowercased version
        selected_lower = [b.strip().lower() for b in selected_brands]
        available_df = available_df[available_df["Brand"].isin(selected_lower)]

    # If nothing remains:
    if available_df.empty:
        print(f"[INFO] No matching brand data in '{csv_path}' after brand filter.")
        return {}

    # Example: add "Strain_Type"
    if "Product" in available_df.columns:
        available_df["Strain_Type"] = available_df["Product"].apply(extract_strain_type)

    # Sort
    sort_cols = []
    if "Category" in available_df.columns:
        sort_cols.append("Category")
    if "Cost" in available_df.columns:
        available_df["Cost"] = pd.to_numeric(available_df["Cost"], errors="coerce")
        sort_cols.append("Cost")
    if "Product" in available_df.columns:
        sort_cols.append("Product")
    if sort_cols:
        available_df.sort_values(by=sort_cols, inplace=True, na_position="last")

    # Drop "Cost"
    if "Cost" in available_df.columns:
        available_df.drop(columns=["Cost"], inplace=True)
    if "Cost" in unavailable_df.columns:
        unavailable_df = unavailable_df.drop(columns=["Cost"])

    # Also normalize brand in the unavailable set
    if "Brand" in unavailable_df.columns and not unavailable_df.empty:
        unavailable_df.loc[:, "Brand"] = unavailable_df["Brand"].astype(str).str.strip().str.lower()

    os.makedirs(out_dir, exist_ok=True)
    base_csv_name = os.path.splitext(os.path.basename(csv_path))[0]

    # Group the *available* portion by brand
    brand_map = {}
    for brand_name_lower, brand_data in available_df.groupby("Brand"):
        # Grab the "Unavailable" rows for that brand
        brand_unavail = pd.DataFrame()
        if not unavailable_df.empty:
            brand_unavail = unavailable_df[unavailable_df["Brand"] == brand_name_lower]

        dt_str = datetime.now().strftime("%m-%d-%Y")
        out_name = f"{base_csv_name}_{brand_name_lower}_{dt_str}.xlsx"
        out_path = os.path.join(out_dir, out_name)

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            brand_data.to_excel(writer, index=False, sheet_name="Available")
            if not brand_unavail.empty:
                brand_unavail.to_excel(writer, index=False, sheet_name="Unavailable")

        advanced_format_excel(out_path)

        if brand_name_lower not in brand_map:
            brand_map[brand_name_lower] = []
        brand_map[brand_name_lower].append(out_path)

        print(f"[INFO] Created {out_path}")

    return brand_map

def upload_brand_reports_to_drive(brand_reports_map):
    """
    brand_reports_map: { brand_name_lower: [list_of_xlsx_paths] }
    1) Create/find top-level "INVENTORY"
    2) Create date subfolder "YYYY-MM-DD"
    3) For each brand, create brand folder (public), upload
    Return: { brand_name_lower: "https://drive.google.com/drive/folders/<id>"}
    """
    drive_svc = drive_authenticate()
    top_id = find_or_create_folder(drive_svc, DRIVE_PARENT_FOLDER_NAME)
    if not top_id:
        print("[ERROR] Could not find/create top-level folder. Aborting.")
        return {}

    date_str = datetime.now().strftime("%Y-%m-%d")
    date_id = find_or_create_folder(drive_svc, date_str, parent_id=top_id)
    if not date_id:
        print("[ERROR] Could not create/find date subfolder. Aborting.")
        return {}

    brand_links = {}
    for brand_lower, xlsx_list in brand_reports_map.items():
        brand_id = find_or_create_folder(drive_svc, brand_lower, parent_id=date_id, make_public=True)
        if not brand_id:
            print(f"[ERROR] Could not create folder for {brand_lower}")
            continue

        for xfile in xlsx_list:
            try:
                upload_file_to_drive(drive_svc, xfile, brand_id)
                print(f"[DRIVE] Uploaded {os.path.basename(xfile)} => {brand_lower}")
            except Exception as e:
                print(f"[ERROR] Uploading {xfile} => {brand_lower}: {e}")

        link = f"https://drive.google.com/drive/folders/{brand_id}"
        brand_links[brand_lower] = link

    return brand_links

# ----------------- THE GUI (with config.txt) -----------------
class BrandInventoryGUI:
    def __init__(self, master):
        self.master = master
        
        master.title("Brand Inventory Uploader")
        master.geometry("800x600")  # Default window size
        master.resizable(True, True)  # Allow resizing
        self.frame = tk.Frame(master)
        self.frame.pack(padx=10, pady=10)
        self.master.configure(bg="#f5f5f5")
        self.frame.configure(bg="#f5f5f5")

        default_font = ("Segoe UI", 11)
        self.master.option_add("*Font", default_font)
        self.master.option_add("*Background", "#f5f5f5")
        self.master.option_add("*Button.Background", "#4CAF50")
        self.master.option_add("*Button.Foreground", "white")
        self.master.option_add("*Button.Font", ("Segoe UI", 10, "bold"))
        # Load config, if present
        init_in, init_out = load_config()

        self.input_dir_var = tk.StringVar(value=init_in if init_in else "")
        self.output_dir_var = tk.StringVar(value=init_out if init_out else "")
        self.emails_var = tk.StringVar()
        
        # Row 1: input folder
        row1 = tk.Frame(self.frame)
        row1.pack(pady=5, fill="x")
        tk.Label(row1, text="Input Folder:").pack(side="left")
        tk.Entry(row1, textvariable=self.input_dir_var, width=50, relief="solid", bd=1).pack(side="left", padx=5)
        tk.Button(row1, text="Browse", command=self.browse_input).pack(side="left")

        # Row 2: output folder
        row2 = tk.Frame(self.frame)
        row2.pack(pady=5, fill="x")
        tk.Label(row2, text="Output Folder:").pack(side="left")
        tk.Entry(row2, textvariable=self.output_dir_var, width=50, relief="solid", bd=1).pack(side="left", padx=5)
        tk.Button(row2, text="Browse", command=self.browse_output).pack(side="left")

        # Row 3: Buttons => get files, load brands
        row3 = tk.Frame(self.frame)
        row3.pack(pady=5, fill="x")
        tk.Button(row3, text="Update Files", command=self.get_files, width=15).pack(side="left", padx=10)
        tk.Button(row3, text="Load Brands", command=self.load_brands, width=15).pack(side="left", padx=10)
                # Row 5: emails
        row5 = tk.Frame(self.frame)
        row5.pack(pady=5, fill="x")
        tk.Label(row5, text="Email(s) (comma-separated):").pack(anchor="w")
        tk.Entry(row5, textvariable=self.emails_var, width=60, relief="solid", bd=1).pack()

        # Row 6: final run
        row6 = tk.Frame(self.frame)
        row6.pack(pady=1)
        tk.Button(row6, text="Generate & Upload & Email", command=self.run_process, width=30).pack(padx=3, pady=5)
        # Row 4: brand listbox
        row4 = tk.Frame(self.frame)
        row4.pack(pady=0, fill="both")
        alpha_sidebar = tk.Frame(row4)
        alpha_sidebar.pack(side="right", padx=5, pady=5)

        letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        cols = 4
        for i, letter in enumerate(letters):
            btn = tk.Button(alpha_sidebar, text=letter, width=3, command=lambda l=letter: self.scroll_to_letter(l))
            btn.grid(row=i % 7, column=i // 7, padx=1, pady=1)  # 7 rows max

        tk.Label(row4, text="Select brand(s):").pack(anchor="w")
        self.brand_listbox = tk.Listbox(row4, selectmode=tk.MULTIPLE, height=18, width=50)
        self.brand_listbox.pack(side="left", fill="both", expand=True)
        scroll = tk.Scrollbar(row4, command=self.brand_listbox.yview)
        scroll.pack(side="right", fill="y")
        self.brand_listbox.config(borderwidth=1, relief="solid", highlightthickness=0, bg="white", selectbackground="#4CAF50", selectforeground="white")
        scroll.config(bg="#f5f5f5", troughcolor="#e0e0e0", borderwidth=0)
        self.brand_listbox.bind("<Key>", self.on_listbox_keypress)

    def show_loading(self, message="Processing..."):
        if hasattr(self, "loading_overlay") and self.loading_overlay.winfo_exists():
            return  # Already shown

        self.loading_overlay = tk.Frame(self.master, bg="#ffffff", bd=2, relief="ridge")
        self.loading_overlay.place(relx=0.25, rely=0.4, relwidth=0.5, relheight=0.2)

        self.loading_label = tk.Label(
            self.loading_overlay,
            text=message,
            font=("Segoe UI", 14, "bold"),
            bg="#ffffff",
            fg="#333333"
        )
        self.loading_label.place(relx=0.5, rely=0.5, anchor="center")

        self.master.update()  # Force GUI redraw

    def hide_loading(self):
        if hasattr(self, "loading_overlay") and self.loading_overlay.winfo_exists():
            self.loading_overlay.destroy()
    def scroll_to_letter(self, letter):
        for i in range(self.brand_listbox.size()):
            item = self.brand_listbox.get(i)
            if item.lower().startswith(letter.lower()):
                self._flash_listbox_item(i)
                break

    def on_listbox_keypress(self, event):
        typed_char = event.char.lower()
        if not typed_char.isalpha():
            return

        # Don't deselect current selection
        for i in range(self.brand_listbox.size()):
            item = self.brand_listbox.get(i)
            if item.lower().startswith(typed_char):
                self._flash_listbox_item(i)
                break
    def _flash_listbox_item(self, index):
        # Save current selection
        current_selection = self.brand_listbox.curselection()

        # Highlight item visually
        self.brand_listbox.selection_clear(0, tk.END)
        self.brand_listbox.selection_set(index)
        self.brand_listbox.activate(index)
        self.brand_listbox.see(index)

        # Delay to restore previous selection
        def restore_selection():
            self.brand_listbox.selection_clear(0, tk.END)
            for idx in current_selection:
                self.brand_listbox.selection_set(idx)

        self.master.after(700, restore_selection)  # 700ms flash
    
    def browse_input(self):
        folder = filedialog.askdirectory()
        if folder:
            self.input_dir_var.set(folder)

    def browse_output(self):
        folder = filedialog.askdirectory()
        if folder:
            self.output_dir_var.set(folder)

    def get_files(self):
        """
        Clears the input folder and calls getCatalog.py to fetch new CSVs.
        Shows a loading screen while running.
        """
        in_dir = self.input_dir_var.get().strip()

        if not in_dir or not os.path.isdir(in_dir):
            messagebox.showerror("Error", "Please choose a valid input folder first.")
            return
        if not os.path.exists("getCatalog.py"):
            messagebox.showwarning("Warning", "No getCatalog.py found in this directory.")
            return

        # ✅ Show loading screen
        self.show_loading("Updating files...")

        try:
            # Clear the input folder
            for file in os.listdir(in_dir):
                file_path = os.path.join(in_dir, file)
                if os.path.isfile(file_path):
                    os.remove(file_path)
            print(f"[INFO] Cleared files in input folder: {in_dir}")

            # Run getCatalog
            subprocess.check_call(["python", "getCatalog.py", in_dir])
            self.hide_loading()
            messagebox.showinfo("Success", "CSV files fetched from getCatalog.py (after clearing input folder).")
        except subprocess.CalledProcessError as e:
            self.hide_loading()
            messagebox.showerror("Error", f"getCatalog.py failed:\n{e}")
        except Exception as e:
            self.hide_loading()
            messagebox.showerror("Error", str(e))

    def load_brands(self):
        """
        Scans CSVs in input folder for a 'Brand' column, collects unique brand names,
        lowercases/strips them, and displays them in the listbox.
        """
        in_dir = self.input_dir_var.get().strip()
        if not in_dir or not os.path.isdir(in_dir):
            messagebox.showerror("Error", "Invalid input folder.")
            return

        self.brand_listbox.delete(0, tk.END)
        self.brand_listbox.configure(state="normal")

        brand_set = set()
        for fn in os.listdir(in_dir):
            if fn.lower().endswith(".csv"):
                path = os.path.join(in_dir, fn)
                try:
                    df = pd.read_csv(path, nrows=50000)
                    if "Brand" in df.columns:
                        # Convert to lower+strip for consistent matching
                        new_brands = (
                            df["Brand"]
                            .dropna()
                            .astype(str)
                            .str.strip()
                            .str.lower()
                            .unique()
                            .tolist()
                        )
                        brand_set.update(new_brands)
                except:
                    pass

        if not brand_set:
            self.brand_listbox.insert(tk.END, "No brands found.")
            self.brand_listbox.configure(state="disabled")
        else:
            sorted_brands = sorted(list(brand_set))
            for b in sorted_brands:
                self.brand_listbox.insert(tk.END, b)

    def run_process(self):
        in_dir = self.input_dir_var.get().strip()
        out_dir = self.output_dir_var.get().strip()
        emails = self.emails_var.get().strip()

        if not (in_dir and out_dir and emails):
            messagebox.showerror("Error", "Need input folder, output folder, and at least one email address.")
            return
        if not os.path.isdir(in_dir):
            messagebox.showerror("Error", f"Invalid input folder: {in_dir}")
            return
        if not os.path.isdir(out_dir):
            messagebox.showerror("Error", f"Invalid output folder: {out_dir}")
            return

        sel_indices = self.brand_listbox.curselection()
        if sel_indices:
            selected_brands = [self.brand_listbox.get(i) for i in sel_indices]
            if "No brands found." in selected_brands:
                selected_brands = []
        else:
            selected_brands = []
            messagebox.showinfo(
                "No Selection",
                "No brands selected. Will process all brand data from the CSVs."
            )

        # 1) For each CSV => generate XLSX (Available + Unavailable)
        all_brand_map = {}
        try:
            for fname in os.listdir(in_dir):
                if fname.lower().endswith(".csv"):
                    path = os.path.join(in_dir, fname)
                    brand_map = generate_brand_reports(path, out_dir, selected_brands)
                    # Merge
                    for b_name, xlsx_list in brand_map.items():
                        if b_name not in all_brand_map:
                            all_brand_map[b_name] = []
                        all_brand_map[b_name].extend(xlsx_list)

            if not all_brand_map:
                messagebox.showinfo("Done", "No XLSX files generated (possibly no matching data).")
                return

            # 2) Upload to Drive => get brand folder links
            brand_links = upload_brand_reports_to_drive(all_brand_map)
            if not brand_links:
                messagebox.showerror("Error", "No folders created on Drive. Aborting email.")
                return

            # 3) Send an email with each brand link
            lines = []
            for brand_lower, link in brand_links.items():
                lines.append(f"<h3>{brand_lower}</h3>")
                lines.append(f"<p><a href='{link}'>{link}</a></p>")

            joined = "\n".join(lines)
            body_html = f"""
            <html>
              <body>
                <p>Hello,</p>
                <p>Here are the public Drive folders (with Available & Unavailable reports) for each brand:</p>
                {joined}
                <p>Anyone with these links can download the XLSX files.</p>
                <p>Regards,<br>Brand Inventory Bot</p>
              </body>
            </html>
            """
            subject = "Brand Inventory Drive Links"
            send_email_with_gmail_html(subject, body_html, emails)

            # 4) Save the chosen input/output folders to config.txt for next run
            save_config(in_dir, out_dir)

            messagebox.showinfo("Success", "All done! Folders uploaded & email sent.")
        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("Error", f"An error occurred:\n{e}")


# ----------------- MAIN -----------------
def main():
    root = tk.Tk()
    app = BrandInventoryGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
